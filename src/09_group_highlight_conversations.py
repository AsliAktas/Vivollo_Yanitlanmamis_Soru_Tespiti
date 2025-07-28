"""
09_group_highlight_conversations.py
-----------------------------------
Groups messages by conversation_id, sorts by timestamp,
colors only user rows (green/manual, blue/new), leaves bots white,
freezes the top header row, and adds filters on sentiment/kategori/intent.

Output: outputs/vivollo_grouped_highlight.xlsx

Usage:
    python src/09_group_highlight_conversations.py
"""
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

def main():
    base = Path(__file__).resolve().parent.parent

    # 1) Load manual IDs
    data_dir = base / "data"
    manual_file = next(data_dir.glob("vivollo_clean*.xlsx"), None)
    if not manual_file:
        raise FileNotFoundError(f"Manual file not found under {data_dir}")
    df_manual = pd.read_excel(manual_file)

    # 2) Load final preview
    final_xl = base / "outputs" / "vivollo_final_clean.xlsx"
    df = pd.read_excel(final_xl, sheet_name="preview")
    if "text" not in df.columns:
        df = pd.read_excel(final_xl, sheet_name="full_text")

    # 3) Select and sort
    cols = [
        "conversation_id","message_id","created_at","sender_type",
        "text","yanitlandi_mi","sentiment","kategori","intent"
    ]
    df = df[cols]
    df["created_at"] = pd.to_datetime(df["created_at"], errors="coerce")
    df = df.sort_values(["conversation_id","created_at"]).reset_index(drop=True)

    # 4) Manual ID set
    manual_ids = set(
        df_manual.loc[df_manual["sender_type"]=="user", "message_id"]
    )

    # 5) Prepare workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "grouped"
    ws.freeze_panes = "A2"

    # Header + filter
    for idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=idx, value=col_name)
        cell.font = Font(bold=True)
    # Determine last column letter for auto_filter
    last_col = chr(64 + len(cols))
    ws.auto_filter.ref = f"A1:{last_col}{len(df)+1}"

    # Fills
    green = PatternFill("solid", fgColor="CCFFCC")
    blue  = PatternFill("solid", fgColor="CCCCFF")

    # 6) Write rows
    for i, row in enumerate(df.itertuples(index=False), start=2):
        is_user = row.sender_type == "user"
        fill = None
        if is_user:
            fill = green if row.message_id in manual_ids else blue

        for j, col in enumerate(cols, start=1):
            val = getattr(row, col)
            # Blank out sentiment/kategori/intent for new user rows
            if is_user and row.message_id not in manual_ids and col in ("sentiment","kategori","intent"):
                val = None

            cell = ws.cell(row=i, column=j, value=val)
            if fill:
                cell.fill = fill

    # 7) Save
    out = base / "outputs" / "vivollo_grouped_highlight.xlsx"
    wb.save(out)
    print(f"✓ Saved grouped & highlighted Excel → {out}")

if __name__ == "__main__":
    main()